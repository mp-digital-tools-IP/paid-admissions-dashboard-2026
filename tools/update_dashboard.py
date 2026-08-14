#!/usr/bin/env python3
"""Build privacy-safe aggregates for the paid admissions dashboard."""

from __future__ import annotations

import argparse
import hashlib
import itertools
import json
import os
import re
import shutil
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


SCHEMA_VERSION = 2
FINANCIAL_TARGET = Decimal("178000000")
REFERENCE_2025 = Decimal("279000000")
EXPECTED_PFHD = 1215
EXPECTED_MARKETING = 5068
FORMS = ("Очная", "Очно-заочная", "Заочная")
LEVELS = ("Бакалавриат и специалитет", "Магистратура", "Аспирантура")
MOSCOW_TZ = ZoneInfo("Europe/Moscow")
CODE_RE = re.compile(r"(?<!\d)(\d{1,2}\.\d{1,2}\.\d{1,2})\.?")
CUBE_ALL = "*"
CUBE_SEPARATOR = "\x1f"

REQUIRED_CONTRACT_COLUMNS = {
    "ФИО",
    "Гражданство физического лица",
    "Конкурсная группа.Уровень подготовки",
    "Факультет",
    "Форма обучения",
    "Код специальности",
    "Конкурсная группа.Направление (специальность)",
    "Приоритет",
    "Состояние заявления",
    "Номер договора",
    "Состояние договора",
    "Размер скидки",
    "Стоимость семестра",
    "Сумма заключенных договоров по 1 семестру",
    "Сумма оплаты",
}

FACULTY_ALIASES = {
    "ФИТ": "Факультет информационных технологий",
    "ФЭУ": "Факультет экономики и управления",
    "ФЭИУ": "Факультет экономики и управления",
    "ФМ": "Факультет машиностроения",
    "ТФ": "Транспортный факультет",
    "ПИШ": 'Передовая инженерная школа технологического лидерства «FDR»',
    "ФУИГХ": "Факультет урбанистики и городского хозяйства",
    "ФИДИЖ": "Институт издательского дела и журналистики",
    "ИИДИЖ": "Институт издательского дела и журналистики",
    "ИГРИК": "Институт графики и искусства книги имени В. А. Фаворского",
    "ФХТИБ": "Факультет химической технологии и биотехнологии",
    "ПФ": "Полиграфический факультет",
    "ПИ": "Полиграфический факультет",
    "ФБК": "Факультет базовых компетенций",
    "ШКОЛА": "Передовая школа инженерного дизайна",
}

PUBLIC_FORBIDDEN_KEYS = {
    "фио", "номердоговора", "contractnumber", "телефон", "email", "почта",
    "паспорт", "снилс", "инн", "адрес", "комментарий",
}


class DataValidationError(RuntimeError):
    pass


@dataclass(frozen=True)
class PlanSheet:
    name: str
    header_row: int
    start_row: int
    faculty_col: int
    code_col: int | None
    name_col: int
    forms: tuple[tuple[str, int, int], ...]
    level: str


PLAN_SHEETS = (
    PlanSheet("Москва", 11, 12, 1, 7, 8, (("Очная", 29, 30), ("Очно-заочная", 47, 48), ("Заочная", 65, 66)), "Бакалавриат и специалитет"),
    PlanSheet("Магистратура", 8, 9, 1, 2, 3, (("Очная", 15, 16), ("Очно-заочная", 29, 30), ("Заочная", 43, 44)), "Магистратура"),
    PlanSheet("Аспирантура", 3, 4, 1, None, 4, (("Очная", 15, 16),), "Аспирантура"),
)


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_code(value: Any) -> str:
    match = CODE_RE.search(clean_text(value))
    return match.group(1) if match else ""


def number(value: Any, *, field: str, row: int, allow_blank: bool = True) -> Decimal:
    if value is None or clean_text(value) == "":
        if allow_blank:
            return Decimal(0)
        raise DataValidationError(f"Строка {row}: поле «{field}» обязательно")
    if isinstance(value, bool):
        raise DataValidationError(f"Строка {row}: поле «{field}» не является числом")
    try:
        result = Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise DataValidationError(f"Строка {row}: некорректное число в поле «{field}»") from exc
    if not result.is_finite():
        raise DataValidationError(f"Строка {row}: некорректное число в поле «{field}»")
    return result


def whole(value: Decimal) -> int:
    return int(value.quantize(Decimal("1")))


def money(value: Decimal) -> int | float:
    rounded = value.quantize(Decimal("0.01"))
    return int(rounded) if rounded == rounded.to_integral() else float(rounded)


def normalize_level(value: Any) -> str:
    text = clean_text(value).casefold()
    if "бакалав" in text or "специал" in text:
        return LEVELS[0]
    if "магистр" in text:
        return LEVELS[1]
    if "аспиран" in text:
        return LEVELS[2]
    return clean_text(value)


def normalize_form(value: Any) -> str:
    text = re.sub(r"\s*-\s*", "-", clean_text(value).casefold().replace("ё", "е"))
    if "очно-заоч" in text:
        return "Очно-заочная"
    if "заоч" in text:
        return "Заочная"
    if "очн" in text:
        return "Очная"
    return clean_text(value)


def citizenship_group(value: Any) -> str:
    text = clean_text(value).casefold()
    return "Россия" if text in {"россия", "российская федерация", "рф"} or "россия" in text else "Иностранное"


def faculty_scopes(value: Any) -> list[str]:
    raw = clean_text(value)
    if not raw:
        return ["Не указано"]
    raw = re.sub(r"\([^)]*\)", "", raw).replace("/школа", "+школа").replace("/Школа", "+школа")
    result: list[str] = []
    for part in [clean_text(item) for item in raw.split("+") if clean_text(item)] or [raw]:
        key = re.sub(r"[^А-ЯA-Z]", "", part.upper().replace("Ё", "Е"))
        mapped = FACULTY_ALIASES.get(key, part)
        if mapped not in result:
            result.append(mapped)
    return result or ["Не указано"]


def display_faculty(scopes: list[str]) -> str:
    return "Совместные программы" if len(scopes) > 1 else scopes[0]


def make_id(*parts: str) -> str:
    return "".join(f"q{char}" for char in hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:14])


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _header_text(sheet, row: int, col: int) -> str:
    return clean_text(sheet.cell(row, col).value).casefold()


def _validate_plan_sheet(sheet, spec: PlanSheet) -> None:
    for _, pfhd_col, marketing_col in spec.forms:
        if "пфхд" not in _header_text(sheet, spec.header_row, pfhd_col):
            raise DataValidationError(f"Лист «{spec.name}»: не найден заголовок ПФХД в колонке {pfhd_col}")
        marketing = _header_text(sheet, spec.header_row, marketing_col)
        if "маркет" not in marketing and not ("платное" in marketing and "пфхд" not in marketing):
            raise DataValidationError(f"Лист «{spec.name}»: не найден маркетинговый заголовок в колонке {marketing_col}")


def parse_plan(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    missing = [spec.name for spec in PLAN_SHEETS if spec.name not in workbook.sheetnames]
    if missing:
        raise DataValidationError(f"В плане отсутствуют листы: {', '.join(missing)}")
    records: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for spec in PLAN_SHEETS:
        sheet = workbook[spec.name]
        _validate_plan_sheet(sheet, spec)
        current_faculty = ""
        for row in range(spec.start_row, sheet.max_row + 1):
            if clean_text(sheet.cell(row, spec.faculty_col).value):
                current_faculty = clean_text(sheet.cell(row, spec.faculty_col).value)
            name_cell = clean_text(sheet.cell(row, spec.name_col).value)
            raw_code = clean_text(sheet.cell(row, spec.code_col).value) if spec.code_col else name_cell
            if spec.code_col and not re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.?", raw_code):
                continue
            code = normalize_code(raw_code)
            if not code or code.endswith(".00"):
                continue
            name = re.sub(r"^\s*\d{1,2}\.\d{1,2}\.\d{1,2}\.?\s*", "", name_cell).strip() or name_cell
            scopes = faculty_scopes(current_faculty)
            for form, pfhd_col, marketing_col in spec.forms:
                pfhd = whole(number(sheet.cell(row, pfhd_col).value, field="ПФХД", row=row))
                marketing = whole(number(sheet.cell(row, marketing_col).value, field="Маркетинговая цель", row=row))
                signature = (spec.level, form, code, name, tuple(scopes), pfhd, marketing)
                if signature in seen:
                    continue
                seen.add(signature)
                records.append({
                    "id": make_id(spec.level, form, code, name, "+".join(scopes)),
                    "level": spec.level, "form": form, "code": code, "directionName": name,
                    "facultyScopes": scopes, "joint": len(scopes) > 1,
                    "pfhdTarget": pfhd, "marketingTarget": marketing,
                })
    totals = {
        "pfhdTarget": sum(item["pfhdTarget"] for item in records),
        "marketingTarget": sum(item["marketingTarget"] for item in records),
    }
    if totals != {"pfhdTarget": EXPECTED_PFHD, "marketingTarget": EXPECTED_MARKETING}:
        raise DataValidationError(
            f"Контрольные суммы детального плана не сошлись: ПФХД={totals['pfhdTarget']} "
            f"(ожидалось {EXPECTED_PFHD}), маркетинг={totals['marketingTarget']} (ожидалось {EXPECTED_MARKETING})"
        )
    source_pfhd = source_marketing = formula_errors = 0
    for sheet_name, rows, columns in (
        ("Москва", (12, 14, 18, 37, 39, 41, 48, 62, 64, 66, 71, 76, 82, 87, 95, 103, 107, 109, 115), ((29, 30), (47, 48), (65, 66))),
        ("Магистратура", (90,), ((15, 16), (29, 30), (43, 44))),
        ("Аспирантура", (45,), ((15, 16),)),
    ):
        sheet = workbook[sheet_name]
        for row in rows:
            for pfhd_col, marketing_col in columns:
                try:
                    source_pfhd += whole(number(sheet.cell(row, pfhd_col).value, field="Групповой ПФХД", row=row))
                    source_marketing += whole(number(sheet.cell(row, marketing_col).value, field="Групповой маркетинг", row=row))
                except DataValidationError:
                    formula_errors += 1
    workbook.close()
    records.sort(key=lambda item: (LEVELS.index(item["level"]), FORMS.index(item["form"]), item["code"], item["directionName"], item["id"]))
    return {
        "schemaVersion": SCHEMA_VERSION,
        "source": {"file": path.name, "sha256": file_sha256(path)},
        "totals": totals,
        "sourceControlTotals": {"pfhdTarget": source_pfhd, "marketingTarget": source_marketing, "formulaErrors": formula_errors},
        "reconciliation": {
            "pfhdDelta": source_pfhd - totals["pfhdTarget"],
            "marketingDelta": source_marketing - totals["marketingTarget"],
            "formulaErrors": formula_errors,
            "status": "Детальные строки являются источником плана; групповые итоги показаны как контроль качества",
        },
        "records": records,
    }


def _headers(sheet) -> dict[str, int]:
    values = [clean_text(value) for value in next(sheet.iter_rows(min_row=6, max_row=6, values_only=True))]
    if len(values) < 24:
        raise DataValidationError(f"Ожидалось 24 колонки, найдено {len(values)}")
    missing = sorted(REQUIRED_CONTRACT_COLUMNS - set(values))
    if missing:
        raise DataValidationError(f"В выгрузке отсутствуют обязательные колонки: {', '.join(missing)}")
    return {name: index for index, name in enumerate(values) if name}


def _empty_metrics() -> dict[str, Any]:
    return {
        "rows": 0, "contracts": 0, "active": 0, "signed": 0, "paid": 0, "discounted": 0,
        "listPriceMin": None, "listPriceMax": None, "listPriceTotal": Decimal(0),
        "contractAmount": Decimal(0), "portfolio": Decimal(0),
        "reportedPayment": Decimal(0), "uniqueContractPayment": Decimal(0),
    }


def _add_reported(metrics: dict[str, Any], payment: Decimal) -> None:
    metrics["rows"] += 1
    metrics["reportedPayment"] += payment


def _add_unique(metrics: dict[str, Any], item: dict[str, Any]) -> None:
    metrics["contracts"] += 1
    metrics["active"] += int(item["active"])
    metrics["signed"] += int(item["signed"])
    metrics["paid"] += int(item["payment"] > 0)
    metrics["discounted"] += int(item["discounted"])
    price = item["listPrice"]
    if price > 0:
        metrics["listPriceMin"] = price if metrics["listPriceMin"] is None else min(metrics["listPriceMin"], price)
        metrics["listPriceMax"] = price if metrics["listPriceMax"] is None else max(metrics["listPriceMax"], price)
        metrics["listPriceTotal"] += price
    metrics["contractAmount"] += item["contractAmount"]
    metrics["portfolio"] += item["contractAmount"] if item["signed"] else Decimal(0)
    metrics["uniqueContractPayment"] += item["payment"]


def _merge_metrics(target: dict[str, Any], source: dict[str, Any]) -> None:
    for key in ("rows", "contracts", "active", "signed", "paid", "discounted"):
        target[key] += source[key]
    for key in ("listPriceTotal", "contractAmount", "portfolio", "reportedPayment", "uniqueContractPayment"):
        target[key] += source[key]
    if source["listPriceMin"] is not None:
        target["listPriceMin"] = source["listPriceMin"] if target["listPriceMin"] is None else min(target["listPriceMin"], source["listPriceMin"])
        target["listPriceMax"] = source["listPriceMax"] if target["listPriceMax"] is None else max(target["listPriceMax"], source["listPriceMax"])


def _finalize_metrics(metrics: dict[str, Any]) -> dict[str, Any]:
    result = dict(metrics)
    for key in ("listPriceMin", "listPriceMax", "listPriceTotal", "contractAmount", "portfolio", "reportedPayment", "uniqueContractPayment"):
        result[key] = None if result[key] is None else money(result[key])
    return result


def _signature(item: dict[str, Any]) -> tuple[Any, ...]:
    return tuple(item[key] for key in (
        "person", "level", "form", "code", "directionName", "scopes", "priority",
        "statementStatus", "contractStatus", "discountSize", "listPrice", "contractAmount", "payment",
    ))


def _cube_key(values: Iterable[str]) -> str:
    return CUBE_SEPARATOR.join(values)


def _build_people_cube(rows: list[dict[str, Any]]) -> tuple[dict[str, int], int]:
    people: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in rows:
        people[item["personKey"]].append(item)
    cube: dict[str, int] = defaultdict(int)
    for person_rows in people.values():
        keys: set[str] = set()
        for item in person_rows:
            faculties = set(item["scopes"])
            faculties.add(display_faculty(list(item["scopes"])))
            for faculty in faculties:
                exact = (
                    item["level"], item["form"], faculty, item["code"], item["citizenship"],
                    item["discount"], str(item["priority"]),
                )
                for mask in itertools.product((False, True), repeat=len(exact)):
                    keys.add(_cube_key(CUBE_ALL if wildcard else value for value, wildcard in zip(exact, mask)))
        for key in keys:
            cube[key] += 1
    return dict(cube), len(people)


def _breakdown(rows: list[dict[str, Any]], field: str) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in rows:
        groups[str(item[field])].append(item)
    result = []
    for value, items in groups.items():
        unique_contracts = {item["contract"]: item for item in items}
        result.append({
            "value": value,
            "rows": len(items),
            "uniqueContracts": len(unique_contracts),
            "uniquePeople": len({item["personKey"] for item in items}),
            "reportedPayment": money(sum((item["payment"] for item in items), Decimal(0))),
            "uniqueContractPayment": money(sum((item["payment"] for item in unique_contracts.values()), Decimal(0))),
            "portfolio": money(sum((item["contractAmount"] for item in unique_contracts.values() if item["signed"]), Decimal(0))),
        })
    def sort_key(item: dict[str, Any]):
        if field == "priority":
            return (int(item["value"]),)
        return (item["value"],)
    return sorted(result, key=sort_key)


def parse_contracts(path: Path, plan: dict[str, Any], snapshot_at: datetime) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Лист_1" not in workbook.sheetnames:
        raise DataValidationError("В выгрузке отсутствует лист «Лист_1»")
    sheet = workbook["Лист_1"]
    headers = _headers(sheet)
    plan_keys = {(item["level"], item["form"], item["code"]) for item in plan["records"]}
    rows: list[dict[str, Any]] = []
    summary_rows = 0
    source_total = Decimal(0)
    invalid_amounts = 0
    for row_number, values in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        get = lambda name: values[headers[name]] if headers[name] < len(values) else None
        contract = clean_text(get("Номер договора"))
        payment = number(get("Сумма оплаты"), field="Сумма оплаты", row=row_number)
        if not contract:
            summary_rows += 1
            source_total += payment
            continue
        discount_size = number(get("Размер скидки"), field="Размер скидки", row=row_number)
        list_price = number(get("Стоимость семестра"), field="Стоимость семестра", row=row_number)
        contract_amount = number(get("Сумма заключенных договоров по 1 семестру"), field="Сумма договора", row=row_number)
        if min(discount_size, list_price, contract_amount, payment) < 0:
            invalid_amounts += 1
        priority = whole(number(get("Приоритет"), field="Приоритет", row=row_number))
        level = normalize_level(get("Конкурсная группа.Уровень подготовки"))
        form = normalize_form(get("Форма обучения"))
        code = normalize_code(get("Код специальности"))
        direction = clean_text(get("Конкурсная группа.Направление (специальность)"))
        scopes = faculty_scopes(get("Факультет"))
        statement_status = clean_text(get("Состояние заявления")) or "Не указано"
        contract_status = clean_text(get("Состояние договора")) or "Не указано"
        person = clean_text(get("ФИО"))
        person_key = person.casefold() if person else f"contract:{contract}"
        item = {
            "person": person, "personKey": person_key, "contract": contract,
            "level": level, "form": form, "code": code, "directionName": direction,
            "scopes": tuple(scopes), "joint": len(scopes) > 1,
            "citizenship": citizenship_group(get("Гражданство физического лица")),
            "priority": priority, "statementStatus": statement_status, "contractStatus": contract_status,
            "discountSize": discount_size, "discounted": discount_size > 0,
            "discount": "Есть скидка" if discount_size > 0 else "Без скидки",
            "listPrice": list_price, "contractAmount": contract_amount, "payment": payment,
            "signed": contract_status == "Подписан",
            "active": statement_status != "Отозвано" and contract_status != "Отменен",
            "matched": (level, form, code) in plan_keys,
        }
        rows.append(item)
    workbook.close()
    if invalid_amounts:
        raise DataValidationError(f"Обнаружены отрицательные суммы: {invalid_amounts}")

    seen: dict[str, tuple[Any, ...]] = {}
    unique_rows: list[dict[str, Any]] = []
    duplicate_rows: list[dict[str, Any]] = []
    conflicting_duplicates = 0
    for item in rows:
        signature = _signature(item)
        if item["contract"] in seen:
            if seen[item["contract"]] != signature:
                conflicting_duplicates += 1
                continue
            duplicate_rows.append(item)
            continue
        seen[item["contract"]] = signature
        unique_rows.append(item)
    if conflicting_duplicates:
        raise DataValidationError(f"Обнаружены конфликтующие строки одного договора: {conflicting_duplicates}")

    segments: dict[tuple[Any, ...], dict[str, Any]] = {}
    for item in rows:
        key = (
            item["level"], item["form"], item["code"], item["directionName"], item["scopes"],
            item["citizenship"], item["discount"], item["priority"], item["statementStatus"], item["contractStatus"], item["matched"],
        )
        segments.setdefault(key, _empty_metrics())
        _add_reported(segments[key], item["payment"])
    for item in unique_rows:
        key = (
            item["level"], item["form"], item["code"], item["directionName"], item["scopes"],
            item["citizenship"], item["discount"], item["priority"], item["statementStatus"], item["contractStatus"], item["matched"],
        )
        _add_unique(segments[key], item)

    totals = _empty_metrics()
    for metrics in segments.values():
        _merge_metrics(totals, metrics)
    people_cube, unique_people = _build_people_cube(rows)
    public_segments = []
    for key, metrics in segments.items():
        level, form, code, direction, scopes, citizenship, discount, priority, statement, contract_status, matched = key
        public_segments.append({
            "id": make_id(level, form, code, direction, "+".join(scopes), citizenship, discount, str(priority), statement, contract_status),
            "level": level, "form": form, "code": code, "directionName": direction,
            "facultyScopes": list(scopes), "joint": len(scopes) > 1,
            "citizenship": citizenship, "discount": discount, "priority": priority,
            "statementStatus": statement, "contractStatus": contract_status, "matchedPlan": matched,
            **_finalize_metrics(metrics),
        })

    finalized = _finalize_metrics(totals)
    finalized.update({
        "uniquePeople": unique_people,
        "financialTarget": money(FINANCIAL_TARGET),
        "reference2025": money(REFERENCE_2025),
        "remainingToTarget": money(max(FINANCIAL_TARGET - Decimal(str(finalized["reportedPayment"])), Decimal(0))),
    })

    p1_rows = [item for item in rows if item["priority"] == 1]
    p1_after_withdrawn = [item for item in p1_rows if item["statementStatus"] != "Отозвано"]
    p1_active_rows = [item for item in p1_after_withdrawn if item["contractStatus"] != "Отменен"]
    p1_seen: set[str] = set()
    p1_unique = []
    p1_duplicates = []
    for item in p1_active_rows:
        if item["contract"] in p1_seen:
            p1_duplicates.append(item)
        else:
            p1_seen.add(item["contract"])
            p1_unique.append(item)
    operational = _empty_metrics()
    for item in p1_active_rows:
        _add_reported(operational, item["payment"])
    for item in p1_unique:
        _add_unique(operational, item)
    unmatched = [item for item in rows if not item["matched"]]
    partial = over = payment_without_signed = discount_mismatch = 0
    for item in unique_rows:
        if item["payment"] > 0 and item["contractAmount"] > 0:
            partial += int(item["payment"] < item["contractAmount"])
            over += int(item["payment"] > item["contractAmount"])
        payment_without_signed += int(item["payment"] > 0 and not item["signed"])
        if item["listPrice"] > 0:
            expected = item["listPrice"] * (Decimal(1) - item["discountSize"] / Decimal(100))
            discount_mismatch += int(abs(expected - item["contractAmount"]) > Decimal("1"))

    reported_payment = sum((item["payment"] for item in rows), Decimal(0))
    unique_payment = sum((item["payment"] for item in unique_rows), Decimal(0))
    other_priority = [item for item in rows if item["priority"] != 1]
    withdrawn_p1 = [item for item in p1_rows if item["statementStatus"] == "Отозвано"]
    cancelled_p1 = [item for item in p1_after_withdrawn if item["contractStatus"] == "Отменен"]
    published_at = datetime.now(MOSCOW_TZ).replace(microsecond=0).isoformat()
    return {
        "schemaVersion": SCHEMA_VERSION,
        "snapshotAt": snapshot_at.isoformat(), "publishedAt": published_at,
        "source": {"file": path.name, "sha256": file_sha256(path)},
        "metrics": finalized,
        "dimensions": {
            "levels": [level for level in LEVELS if any(item["level"] == level for item in public_segments)],
            "forms": [form for form in FORMS if any(item["form"] == form for item in public_segments)],
            "faculties": sorted({faculty for item in public_segments for faculty in item["facultyScopes"]}),
            "directions": sorted(({"code": item["code"], "name": item["directionName"]} for item in public_segments), key=lambda item: (item["code"], item["name"])),
            "citizenships": ["Россия", "Иностранное"],
            "discounts": ["Без скидки", "Есть скидка"],
            "priorities": sorted({item["priority"] for item in public_segments}),
        },
        "segments": sorted(public_segments, key=lambda item: (LEVELS.index(item["level"]) if item["level"] in LEVELS else 99, FORMS.index(item["form"]) if item["form"] in FORMS else 99, item["code"], item["priority"], item["citizenship"])),
        "peopleCube": {"all": CUBE_ALL, "separator": CUBE_SEPARATOR, "counts": people_cube},
        "breakdowns": {
            "priorities": _breakdown(rows, "priority"),
            "citizenship": _breakdown(rows, "citizenship"),
            "contractStatuses": _breakdown(rows, "contractStatus"),
            "statementStatuses": _breakdown(rows, "statementStatus"),
        },
        "quality": {
            "sourceRows": len(rows) + summary_rows, "detailRows": len(rows), "summaryRows": summary_rows,
            "sourceSummaryPayment": money(source_total), "sourceSummaryDelta": money(source_total - reported_payment),
            "duplicateContracts": len(duplicate_rows), "duplicatePayment": money(reported_payment - unique_payment),
            "conflictingDuplicates": conflicting_duplicates,
            "unmatchedDirections": len({(item["level"], item["form"], item["code"], item["directionName"]) for item in unmatched}),
            "unmatchedRows": len(unmatched), "unmatchedPayment": money(sum((item["payment"] for item in unmatched), Decimal(0))),
            "matchedDirectionCodes": len({item["code"] for item in rows if item["matched"]}),
            "discountFormulaMismatches": discount_mismatch, "partialPayments": partial,
            "overpayments": over, "paymentsWithoutSignedStatus": payment_without_signed,
            "planReconciliation": plan["reconciliation"],
            "reconciliation": {
                "reported": {"rows": len(rows), "payment": money(reported_payment)},
                "uniqueContracts": {"contracts": len(unique_rows), "payment": money(unique_payment)},
                "operationalPriorityOne": {**_finalize_metrics(operational), "uniquePeople": len({item["personKey"] for item in p1_unique})},
                "steps": [
                    {"label": "Другие приоритеты", "rows": len(other_priority), "payment": money(sum((item["payment"] for item in other_priority), Decimal(0)))},
                    {"label": "Отозвано, приоритет 1", "rows": len(withdrawn_p1), "payment": money(sum((item["payment"] for item in withdrawn_p1), Decimal(0)))},
                    {"label": "Отменено, приоритет 1", "rows": len(cancelled_p1), "payment": money(sum((item["payment"] for item in cancelled_p1), Decimal(0)))},
                    {"label": "Точные дубли, приоритет 1", "rows": len(p1_duplicates), "payment": money(sum((item["payment"] for item in p1_duplicates), Decimal(0)))},
                ],
            },
        },
    }


def _plan_fingerprint(plan: dict[str, Any]) -> str:
    body = {"totals": plan["totals"], "records": plan["records"]}
    return hashlib.sha256(json.dumps(body, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()


def _read_json(path: Path, default: Any) -> Any:
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else default


def _public_privacy_check(payloads: Iterable[tuple[str, Any]]) -> None:
    email_re = re.compile(r"[\w.+-]+@[\w.-]+\.[A-Za-zА-Яа-я]{2,}")
    phone_re = re.compile(r"(?<!\d)(?:\+7|8)[\s()\-]*\d{3}[\s()\-]*\d{3}[\s\-]*\d{2}[\s\-]*\d{2}(?!\d)")
    def walk(value: Any, path: str) -> None:
        if isinstance(value, dict):
            for key, child in value.items():
                normalized = re.sub(r"[^a-zа-я0-9]", "", str(key).casefold())
                if normalized in PUBLIC_FORBIDDEN_KEYS:
                    raise DataValidationError(f"Запрещённое поле в публичных данных: {path}.{key}")
                walk(child, f"{path}.{key}")
        elif isinstance(value, list):
            for index, child in enumerate(value):
                walk(child, f"{path}[{index}]")
        elif isinstance(value, str) and not path.endswith(".id") and (email_re.search(value) or phone_re.search(value)):
            raise DataValidationError(f"Возможные персональные данные в публичном поле: {path}")
    for name, payload in payloads:
        walk(payload, name)


def build_history(existing: dict[str, Any], snapshot: dict[str, Any]) -> dict[str, Any]:
    date = snapshot["snapshotAt"][:10]
    points = [point for point in existing.get("points", []) if point.get("date") != date]
    points.append({"date": date, "snapshotAt": snapshot["snapshotAt"], "publishedAt": snapshot["publishedAt"], "metrics": snapshot["metrics"]})
    return {"schemaVersion": SCHEMA_VERSION, "points": sorted(points, key=lambda point: point["date"])}


def publish_atomic(public_data: Path, payloads: dict[Path, Any]) -> None:
    public_data.mkdir(parents=True, exist_ok=True)
    stage = Path(tempfile.mkdtemp(prefix=".dashboard-stage-", dir=public_data.parent))
    backups: dict[Path, bytes | None] = {}
    try:
        for relative, payload in payloads.items():
            destination = stage / relative
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            json.loads(destination.read_text(encoding="utf-8"))
        for relative in payloads:
            target = public_data / relative
            backups[target] = target.read_bytes() if target.exists() else None
        try:
            for relative in payloads:
                target = public_data / relative
                target.parent.mkdir(parents=True, exist_ok=True)
                os.replace(stage / relative, target)
        except Exception:
            for target, content in backups.items():
                if content is None:
                    target.unlink(missing_ok=True)
                else:
                    target.write_bytes(content)
            raise
    finally:
        shutil.rmtree(stage, ignore_errors=True)


def parse_snapshot_at(value: str) -> datetime:
    try:
        result = datetime.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Используйте ISO-время, например 2026-08-13T22:57:54+03:00") from exc
    if result.tzinfo is None:
        raise argparse.ArgumentTypeError("Время среза должно содержать часовой пояс")
    return result.astimezone(MOSCOW_TZ).replace(microsecond=0)


def update(plan_path: Path, contracts_path: Path, snapshot_at: datetime, output_dir: Path, allow_plan_change: bool = False) -> dict[str, Any]:
    if not plan_path.is_file() or not contracts_path.is_file():
        raise DataValidationError("Файл плана или выгрузки договоров не найден")
    plan = parse_plan(plan_path)
    existing_plan_path = output_dir / "plan.json"
    if existing_plan_path.exists() and not allow_plan_change:
        existing_plan = _read_json(existing_plan_path, {})
        if _plan_fingerprint(existing_plan) != _plan_fingerprint(plan):
            raise DataValidationError("План ПФХД изменился. Повторите с --allow-plan-change после явного согласования")
    snapshot = parse_contracts(contracts_path, plan, snapshot_at)
    history = build_history(_read_json(output_dir / "history.json", {}), snapshot)
    payloads = {
        Path("plan.json"): plan, Path("current.json"): snapshot, Path("history.json"): history,
        Path("snapshots") / f"{snapshot_at.date().isoformat()}.json": snapshot,
    }
    _public_privacy_check((str(path), payload) for path, payload in payloads.items())
    publish_atomic(output_dir, payloads)
    return {"plan": plan["totals"], "metrics": snapshot["metrics"], "quality": snapshot["quality"], "snapshotAt": snapshot["snapshotAt"]}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--plan", required=True, type=Path)
    parser.add_argument("--contracts", required=True, type=Path)
    parser.add_argument("--snapshot-at", required=True, type=parse_snapshot_at)
    parser.add_argument("--output", type=Path, default=Path("public/data"))
    parser.add_argument("--allow-plan-change", action="store_true")
    args = parser.parse_args()
    try:
        report = update(args.plan, args.contracts, args.snapshot_at, args.output, args.allow_plan_change)
    except DataValidationError as exc:
        parser.exit(2, f"ОШИБКА ПРОВЕРКИ: {exc}\n")
    metrics = report["metrics"]
    quality = report["quality"]
    print(
        "Срез обновлён:", report["snapshotAt"],
        f"| по выгрузке {metrics['reportedPayment']} руб.",
        f"| уникальные договоры {metrics['contracts']}",
        f"| уникальные люди {metrics['uniquePeople']}",
        f"| дубли {quality['duplicateContracts']}",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
