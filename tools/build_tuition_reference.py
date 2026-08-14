#!/usr/bin/env python3
"""Build a privacy-safe annual tuition reference from local source files."""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from collections import Counter, defaultdict
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook

from update_dashboard import (
    DataValidationError,
    FORMS,
    LEVELS,
    _headers,
    clean_text,
    faculty_scopes,
    make_id,
    normalize_code,
    normalize_form,
    normalize_level,
    number,
    parse_plan,
    whole,
)


ORDER_LABELS = {
    "бакалавриат": LEVELS[0],
    "специалитет": LEVELS[0],
    "магистратура": LEVELS[1],
    "аспирантура": LEVELS[2],
}


def _money_cell(value: Any) -> int:
    text = clean_text(value).replace("руб.", "").replace("руб", "")
    return whole(number(text, field="Стоимость обучения", row=0, allow_blank=False))


def parse_order(path: Path) -> list[dict[str, Any]]:
    """Read annual first-year tuition rows from the official order table."""
    rows: list[dict[str, Any]] = []
    level = ""
    with pdfplumber.open(path) as document:
        for page in document.pages:
            for table in page.extract_tables():
                for cells in table:
                    if not cells:
                        continue
                    first = clean_text(cells[0])
                    label = first.casefold()
                    if label in ORDER_LABELS:
                        level = ORDER_LABELS[label]
                        continue
                    code = normalize_code(first)
                    if not level or not code or len(cells) < 4:
                        continue
                    form = normalize_form(cells[2])
                    if form not in FORMS:
                        continue
                    rows.append({
                        "level": level,
                        "form": form,
                        "code": code,
                        "directionName": clean_text(cells[1]),
                        "annualCost": _money_cell(cells[3]),
                    })
    if not rows:
        raise DataValidationError("В приказе не найдены строки стоимости обучения")
    return rows


def parse_pfhd_costs(path: Path) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    """Read the secondary annual cost field available in the PFHD plan."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Москва"]
    result: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in range(12, sheet.max_row + 1):
        code = normalize_code(sheet.cell(row, 7).value)
        raw_cost = sheet.cell(row, 17).value
        if not code or raw_cost in (None, ""):
            continue
        try:
            annual_cost = whole(number(raw_cost, field="Стоимость ОП", row=row, allow_blank=False))
        except DataValidationError:
            continue
        if annual_cost <= 0:
            continue
        result[(LEVELS[0], "Очная", code)].append({
            "directionName": clean_text(sheet.cell(row, 8).value),
            "annualCost": annual_cost,
        })
    workbook.close()
    return result


def parse_contract_costs(path: Path) -> tuple[dict[tuple[str, str, str], Counter[int]], dict[tuple[str, str, str], dict[str, Any]]]:
    """Aggregate semester costs without exposing contract or person identifiers."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Лист_1" not in workbook.sheetnames:
        raise DataValidationError("В выгрузке отсутствует лист «Лист_1»")
    sheet = workbook["Лист_1"]
    headers = _headers(sheet)
    costs: dict[tuple[str, str, str], Counter[int]] = defaultdict(Counter)
    metadata: dict[tuple[str, str, str], dict[str, Any]] = {}
    seen_contracts: set[str] = set()
    for row_number, values in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
        get = lambda name: values[headers[name]] if headers[name] < len(values) else None
        contract = clean_text(get("Номер договора"))
        if not contract or contract in seen_contracts:
            continue
        seen_contracts.add(contract)
        semester_cost = number(get("Стоимость семестра"), field="Стоимость семестра", row=row_number)
        if semester_cost <= 0:
            continue
        key = (
            normalize_level(get("Конкурсная группа.Уровень подготовки")),
            normalize_form(get("Форма обучения")),
            normalize_code(get("Код специальности")),
        )
        if key[0] not in LEVELS or key[1] not in FORMS or not key[2]:
            continue
        annual_cost = whole(semester_cost * Decimal(2))
        costs[key][annual_cost] += 1
        current = metadata.setdefault(key, {"names": Counter(), "scopes": set()})
        current["names"][clean_text(get("Конкурсная группа.Направление (специальность)"))] += 1
        current["scopes"].update(faculty_scopes(get("Факультет")))
    workbook.close()
    return costs, metadata


def _name_key(value: str) -> str:
    return re.sub(r"[^a-zа-я0-9]+", " ", value.casefold().replace("ё", "е")).strip()


def _best_named_cost(candidates: list[dict[str, Any]], direction_name: str) -> int:
    target = _name_key(direction_name)
    target_english = "english" in target

    def score(candidate: dict[str, Any]) -> tuple[int, float, int]:
        name = _name_key(candidate["directionName"])
        english_match = int(("english" in name) == target_english)
        return (english_match, SequenceMatcher(None, target, name).ratio(), -len(name))

    return int(max(candidates, key=score)["annualCost"])


def _mode(counter: Counter[int]) -> int:
    return sorted(counter.items(), key=lambda item: (-item[1], -item[0]))[0][0]


def _collapse(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for record in records:
        key = (
            record["level"], record["form"], record["code"], record["directionName"],
            record["annualCost"], record["sourceKind"],
        )
        current = grouped.setdefault(key, {**record, "facultyScopes": []})
        for scope in record["facultyScopes"]:
            if scope not in current["facultyScopes"]:
                current["facultyScopes"].append(scope)
    result = []
    for record in grouped.values():
        record["joint"] = len(record["facultyScopes"]) > 1
        record["id"] = make_id(
            record["level"], record["form"], record["code"], record["directionName"],
            str(record["annualCost"]), record["sourceKind"],
        )
        result.append(record)
    result.sort(key=lambda item: (
        LEVELS.index(item["level"]), FORMS.index(item["form"]), item["code"], item["directionName"], item["annualCost"],
    ))
    return result


def build_reference(order_path: Path, plan_path: Path, contracts_path: Path) -> dict[str, Any]:
    plan = parse_plan(plan_path)
    order_rows = parse_order(order_path)
    pfhd_costs = parse_pfhd_costs(plan_path)
    contract_costs, contract_metadata = parse_contract_costs(contracts_path)

    plan_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for item in plan["records"]:
        plan_by_key[(item["level"], item["form"], item["code"])].append(item)

    records: list[dict[str, Any]] = []
    represented_keys: set[tuple[str, str, str]] = set()
    missing: list[dict[str, Any]] = []

    # Preserve every official row: the order is the primary source.
    for order_item in order_rows:
        key = (order_item["level"], order_item["form"], order_item["code"])
        plan_candidates = plan_by_key.get(key, [])
        if not plan_candidates and order_item["level"] == LEVELS[1]:
            wrong_bachelor_code = order_item["code"].replace(".04.", ".03.")
            plan_candidates = plan_by_key.get((order_item["level"], order_item["form"], wrong_bachelor_code), [])
        scopes: list[str] = []
        if plan_candidates:
            best_plan = max(
                plan_candidates,
                key=lambda item: SequenceMatcher(
                    None, _name_key(order_item["directionName"]), _name_key(item["directionName"])
                ).ratio(),
            )
            scopes = list(best_plan["facultyScopes"])
        elif key in contract_metadata:
            scopes = sorted(contract_metadata[key]["scopes"])
        records.append({
            "level": order_item["level"], "form": order_item["form"], "code": order_item["code"],
            "directionName": order_item["directionName"], "facultyScopes": scopes or ["Не указано"],
            "annualCost": order_item["annualCost"], "sourceKind": "order",
            "sourceLabel": "Приказ № 392-ОД от 08.04.2026",
        })
        represented_keys.add(key)

    # Add only active planned rows for which the order has no level/form/code row.
    for item in plan["records"]:
        key = (item["level"], item["form"], item["code"])
        corrected_key = key
        if item["level"] == LEVELS[1] and ".03." in item["code"]:
            corrected_key = (item["level"], item["form"], item["code"].replace(".03.", ".04."))
        if key in represented_keys or corrected_key in represented_keys:
            continue
        has_activity = bool(item["pfhdTarget"] or item["marketingTarget"] or key in pfhd_costs or key in contract_costs)
        if not has_activity:
            continue
        annual_cost = 0
        source_kind = ""
        source_label = ""
        if key in pfhd_costs:
            annual_cost = _best_named_cost(pfhd_costs[key], item["directionName"])
            source_kind = "pfhd"
            source_label = "План ПФХД"
        elif key in contract_costs:
            annual_cost = _mode(contract_costs[key])
            source_kind = "contracts"
            source_label = "Стоимость семестра × 2"
        if not annual_cost:
            missing.append({
                "level": item["level"], "form": item["form"], "code": item["code"],
                "directionName": item["directionName"],
            })
            continue
        represented_keys.add(key)
        records.append({
            "level": item["level"], "form": item["form"], "code": item["code"],
            "directionName": item["directionName"], "facultyScopes": list(item["facultyScopes"]),
            "annualCost": annual_cost, "sourceKind": source_kind, "sourceLabel": source_label,
        })

    for key, counter in contract_costs.items():
        if key in represented_keys:
            continue
        meta = contract_metadata[key]
        records.append({
            "level": key[0], "form": key[1], "code": key[2],
            "directionName": meta["names"].most_common(1)[0][0] or "Без названия",
            "facultyScopes": sorted(meta["scopes"]), "annualCost": _mode(counter),
            "sourceKind": "contracts", "sourceLabel": "Стоимость семестра × 2",
        })

    records = _collapse(records)
    source_counts = Counter(item["sourceKind"] for item in records)
    return {
        "schemaVersion": 1,
        "basis": "Стоимость обучения за первый год",
        "method": "Приказ → ПФХД → стоимость семестра × 2",
        "order": {"number": "392-ОД", "date": "2026-04-08", "academicYear": "2026/2027"},
        "coverage": {
            "records": len(records), "orderEntries": len(order_rows),
            "fromOrder": source_counts["order"], "fromPfhd": source_counts["pfhd"],
            "fromSemesterCost": source_counts["contracts"], "missingActivePlanRows": len(missing),
        },
        "missing": missing,
        "records": records,
    }

def write_json_atomic(payload: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".tmp", dir=output_path.parent, delete=False) as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
        temporary = Path(handle.name)
    os.replace(temporary, output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--order", required=True, type=Path)
    parser.add_argument("--plan", required=True, type=Path)
    parser.add_argument("--contracts", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    payload = build_reference(args.order, args.plan, args.contracts)
    write_json_atomic(payload, args.output)
    print(json.dumps(payload["coverage"], ensure_ascii=False))


if __name__ == "__main__":
    main()
